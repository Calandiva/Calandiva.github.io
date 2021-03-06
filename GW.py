from openpyxl import load_workbook, Workbook, drawing
from selenium import webdriver
import time
from selenium.webdriver.common.action_chains import ActionChains
import urllib.request
from urllib import request
from requests import get
import cv2
import pytesseract
import numpy as np
from PIL import Image
import shutil, sys, os

################################################################################################
################################################################################################
"""
소개
일일업무보고 내 시간별캡처 파일을 자동으로 확인해주는 프로그램입니다.
확인한 내용을 토대로 각 일일업무보고에 자동으로 댓글 작성이 가능합니다.
일일업무보고에 첨부되는 사진 파일은 PNG 파일이어야 하며, 우측 하단에 워터마크가 표시되어야 합니다.
정상으로 간주하는 시간 범위는 정시 기준 10분 내외 입니다.
워터마크 상의 날짜는 고려하지 않습니다. 다만 출력된 엑셀 파일에서는 확인할 수 있습니다.
"""
################################################################################################
################################################################################################
"""
환경
1. Python 설치 (3.7 권장) - ttps://www.python.org/downloads/
2. Python 환경변수 지정 - 방법 https://wxmin.tistory.com/121
3. tesseract 설치(latest installers) - https://github.com/UB-Mannheim/tesseract/wiki
   (경로 C:\Program Files\Tesseract-OCR\tesseract)
4. ChromeDriver 다운로드 (설치된 크롬과 동일 버전) - https://chromedriver.chromium.org/downloads
   (C:/Users\Chromedriver.exe 위치시키기)
5. 필수 모듈 설치 (CMD에 아래 명령어 하나씩 입력)
pip install opencv-python
pip install selenium
pip install openpyxl
pip install pytesseract
pip install pillow
pip install numpy
"""
################################################################################################
################################################################################################
"""
1. 해당 문서 본문을 복사하여 로컬에 py파일로 저장
2. 담당 재택근로자 정보 수정 (line.186) 
3. 그룹웨어 ID, PW 입력 (line.55)
4. 댓글 작성 여부 확인 (line.59 - "Y"인 경우 댓글 자동 작성, 그 외 pass)

"""
################################################################################################
################################################################################################

# ID, PW 정보 작성
userId = "hije"
userPw = "pwpwpw"

# 댓글작성여부 Y or Anykey
det = 'Y'

# 현재 시간 불러오기
now = time.localtime()
s = "%04d_%02d_%02d_%02d_%02d_%02d" % (now.tm_year, now.tm_mon, now.tm_mday, now.tm_hour, now.tm_min, now.tm_sec)

# 파일명 생성
file_name = 'ummubogo' + '.xlsx'

try:
    # 기초데이터임포트
    wb = load_workbook(filename=file_name, data_only=True)
    ws1 = wb['Sheet']

    url_list = []

    for eu in range(2, ws1.max_row + 1):
        # 셀이 참조할 숫자
        numb = ws1.max_row

        oldu = 'L%s' % eu
        ou_v = ws1[oldu].value
        url_list.append(ou_v)
        print(ou_v)

    print('기존 데이터 : ', ws1.max_row)

except:
    url_list = []
    numb = 2
    print('파일 존재하지 않음. 해당 위치에 신규 생성')

    # 워크북 생성
    wb = Workbook()
    # 워크시트 생성
    ws1 = wb.active

    # 워크시트 입력
    ws1['A%s' % 1] = "순번"
    ws1['B%s' % 1] = "이름"
    ws1['C%s' % 1] = "제목"
    ws1['D%s' % 1] = "날짜"
    ws1['E%s' % 1] = "파일수"
    ws1['F%s' % 1] = '누락 사진 수'
    ws1['G%s' % 1] = '누락 시간'
    ws1['H%s' % 1] = "보고사항"
    ws1['I%s' % 1] = "특이사항"
    ws1['J%s' % 1] = "파일명"
    ws1['K%s' % 1] = "OCR"
    ws1['L%s' % 1] = "주소"

# 웹드라이버 옵션
options = webdriver.ChromeOptions()
options.add_argument('headless')
options.add_argument('window-size=1920x1080')
options.add_argument("disable-gpu")
# options.add_argument("--disable-gpu")

# 브라우저 실행 - 웹드라이버 경로 설정 필요
driver = webdriver.Chrome("C:/Users\Chromedriver.exe")

# 주소이동 - 그룹웨어(로그인 목적)
driver.get("https://gw.kcopa.or.kr/gw")

# 경고창 확인
try:
    # 얼러트 종료
    time.sleep(1)
    alert = driver.switch_to.alert
    alert.accept()
except:
    pass

handles = driver.window_handles
b_window = driver.window_handles[0]

# 아이디 입력
id = driver.find_element_by_name("id")
id.send_keys(userId)

# 패스워드 입력
pw = driver.find_element_by_id("userPw")
pw.send_keys(userPw)

# 로그인 버튼 클릭
driver.find_element_by_class_name("log_btn").click()

# 5초 대기
time.sleep(3)

try:
    # 창 번호 찾기
    handles = driver.window_handles

    if len(handles) == 1:
        b_window = driver.window_handles[0]
        time.sleep(1)
    elif len(handles) == 2:
        a_window = driver.window_handles[1]
        b_window = driver.window_handles[0]
        time.sleep(1)
        # 두 번째 창으로 스위치
        driver.switch_to.window(a_window)
        time.sleep(1)
        driver.close()
    elif len(handles) == 3:
        c_window = driver.window_handles[2]
        a_window = driver.window_handles[1]
        b_window = driver.window_handles[0]
        time.sleep(2)
        # 세 번째 창으로 스위치
        driver.switch_to.window(c_window)
        time.sleep(1)
        driver.close()
        time.sleep(1)
        # 두 번째 창으로 스위치
        driver.switch_to.window(a_window)
        time.sleep(1)
        driver.close()

except:
    print('윈도우관련에러')
    pass

# 재택근로자 이름 및 근무조, 근무요일
jetek_info = {
    '김동수': ['야간', '월화수목일'],
    '김지원': ['심야', '월화수목일'],
    '남미연': ['주간', '월화수금토'],
    '송영민': ['야간', '월화목금토'],
    '김가영': ['야간', '월화수목토'],
    '김은미': ['주간', '월화수금일'],
    '김현정': ['심야', '월화수목일'],
    '노영은': ['주간', '월화수목일'],
    '정민도': ['심야', '월화목금일'],
    '김재필': ['주간', '월화수목일'],
    '박순옥': ['야간', '화수목금토​'],
    '어혜란': ['야간', '월화수목일'],
    '권미정': ['주간', '월화수목일'],
    '김영선': ['주간', '화수목금토​'],
    '김종선': ['주간', '월화목금일'],
    '남유정': ['야간', '월화수금일'],
    '박금배': ['야간', '월화수목토'],
    '정성월': ['야간', '월화목금일'],
    '김기정': ['심야', '화수목금토​'],
    '김수경': ['야간', '월화수목일'],
    '김가람': ['주간', '화수목금토​'],
    '김옥현': ['심야', '월화수목토'],
    '김진영': ['주간', '화수목금토​'],
    '노경화': ['주간', '화수목금토​'],
    '박민영': ['주간', '월화수금토'],
    '하오잉잉': ['야간', '월화수목일'],
    '이상훈': ['심야', '월화수목일'],
    '마하문': ['야간', '월화수목일'],
    '김영호': ['심야', '월화수목토'],
    '임선정': ['야간', '월화수금일'],
    '배재은': ['주간', '월화수목금'],
    '이신덕': ['주간', '화수목금토'],
    '이은영': ['심야', '화수목금토'],
    '이현욱': ['주간', '월화수목금'],
}

# 테서렉트 경로 설정
pytesseract.pytesseract.tesseract_cmd = r'C:\Program Files\Tesseract-OCR\tesseract'


def go_bogo():
    # 메인윈도우로 이동
    driver.switch_to.window(b_window)

    time.sleep(2)

    # 업무보고탭 이동
    driver.get("http://gw.kcopa.or.kr/schedule/Views/Common/report/reportCheckList?menu_no=304020000")

    # 상세검색 버튼 클릭
    # driver.find_element_by_id("all_menu_btn").click()
    # time.sleep(2)

    # 시작달력버튼 클릭
    # driver.find_element_by_xpath("""/html/body/div[1]/div[3]/dl/dd[1]/div/div/div[1]/input[2]""").click()
    # time.sleep(2)

    # 달넘기기(왼쪽)
    # driver.find_element_by_xpath("""/html/body/div[4]/div/div/div[1]/div/div[1]/a[2]""").click()
    # time.sleep(2)
    # driver.find_element_by_xpath("""/html/body/div[4]/div/div/div[1]/div/div[1]/a[2]""").click()
    # time.sleep(2)

    # 달넘기기(오른쪽)
    # driver.find_element_by_xpath("""/html/body/div[4]/div/div/div[1]/div/div[1]/a[3]""").click()
    # time.sleep(2)

    # 일클릭(1번줄4번째, 7월1일 수요일)
    # driver.find_element_by_xpath("""/html/body/div[4]/div/div/div[1]/div/div[2]/table/tr[2]/td[4]/a""").click()
    # time.sleep(2)

    # 일클릭( 8월31일 토요일)
    # driver.find_element_by_xpath("""/html/body/div[4]/div/div/div[1]/div/div[2]/table/tr[7]/td[2]/a""").click()
    # time.sleep(2)

    # 일클릭(1번줄7번째, 8월6일 토요일)
    # driver.find_element_by_xpath("""/html/body/div[4]/div/div/div[1]/div/div[2]/table/tr[3]/td[5]/a""").click()
    # time.sleep(2)

    # 일클릭(3번줄3번째-14일)
    # driver.find_element_by_xpath("""/html/body/div[4]/div/div/div[1]/div/div[2]/table/tr[4]/td[3]/a""").click()
    # time.sleep(2)

    # 일클릭
    # driver.find_element_by_xpath("""/html/body/div[4]/div/div/div[1]/div/div[2]/table/tr[2]/td[4]""").click()
    # time.sleep(2)

    # 확인
    # driver.find_element_by_class_name("botBtn").click()
    # time.sleep(2)

    # 종료달력
    # driver.find_element_by_xpath("""/html/body/div[1]/div[3]/dl/dd[1]/div/div/div[3]/input[2]""").click()
    # time.sleep(2)

    # 달넘기기
    # driver.find_element_by_xpath("""/html/body/div[4]/div/div/div[2]/div/div[1]/a[2]/span""").click()
    # time.sleep(2)

    # 달넘기기(좌로) 8월전체
    # driver.find_element_by_xpath("""/html/body/div[4]/div/div/div[2]/div/div[1]/a[2]/span""").click()
    # time.sleep(2)

    # 우측달력일클릭(8.31) 8월전체
    # driver.find_element_by_xpath("""/html/body/div[4]/div/div/div[2]/div/div[2]/table/tr[7]/td[2]/a""").click()
    # time.sleep(1)

    # 우측달력일클릭(7.10)
    # driver.find_element_by_xpath("""/html/body/div[4]/div/div/div[2]/div/div[2]/table/tr[3]/td[6]/a""").click()
    # time.sleep(1)

    # 우측달력일클릭(9.1)
    # driver.find_element_by_xpath("""/html/body/div[4]/div/div/div[2]/div/div[2]/table/tr[2]/td[3]/a""").click()
    # time.sleep(1)

    # 확인
    # driver.find_element_by_class_name("botBtn").click()
    # time.sleep(2)

    # 검색버튼 클릭
    # driver.find_element_by_class_name("submit").click()
    # 대기
    time.sleep(2)


def clickman():
    try:
        ia = 1

        # 페이지 (최대 10페이지 한정)
        while ia < 11:

            i = 1

            # 창 컨트롤 - 나중에 이 변수 쓸 것임
            handles = driver.window_handles

            b_window = driver.window_handles[0]

            # 메인화면 포커스
            driver.switch_to.window(b_window)

            # 페이지번호(ia) 클릭
            driver.find_element_by_xpath(
                """/html/body/div[1]/div[4]/div[2]/div[2]/div[1]/ol/li[""" + str(ia) + """]""").click()
            time.sleep(2)

            # 게시물 (20개 기준)
            while i < 21:
                # 메인화면 포커스
                driver.switch_to.window(b_window)
                # 클릭 대상 게시물 초점(i번째)
                one = driver.find_element_by_xpath(
                    "/html/body/div[1]/div[4]/div[2]/div[1]/div[2]/table/tbody/tr[" + str(i) + "]/td[4]")
                # actionChains 변수
                actionChains = ActionChains(driver)
                # actionChains 더블클릭 할당
                actionChains.double_click(one).perform()
                # 대기
                time.sleep(1)

                # 윈도우 2개 이상인 경우임
                b_window = driver.window_handles[0]
                a_window = driver.window_handles[1]

                # 두 번째 창으로 스위치
                driver.switch_to.window(a_window)

                # 현재 URL - cu
                cu = driver.current_url

                if cu in url_list:
                    print('이미 존재함' + cu)
                    driver.switch_to.window(a_window)
                    driver.close()
                    driver.switch_to.window(b_window)
                else:
                    print('카픽돌입')
                    # 사진 관련 함수 실행
                    count_pic()
                    # 메인화면 포커스
                    driver.switch_to.window(b_window)

                if i == 20 and ia == 10:
                    driver.find_element_by_xpath(
                        """/html/body/div[1]/div[4]/div[2]/div[2]/div[1]/span[3]""").click()
                    ia = 0
                else:
                    i = i + 1
                    pass
            ia = ia + 1
    except:
        print('에러')
        pass


def imgcrop(fileid, filenm):
    # 변수 tess를 글로벌로
    global tess

    # 여기서 에러나면 100% 테서렉트 에러임
    try:
        # 다운로드 URL 지정
        img_url1 = "http://gw.kcopa.or.kr/gw/cmm/file/fileDownloadProc.do?fileId=" + fileid
        # 다운로드 URL로 이동, 크롬 자동다운로드 - 폴더 디폴트
        driver.get(img_url1)
        # 대기
        time.sleep(2)

        try:
            # 사진 파일 경로 설정 - 다운로드 폴더 내 파일명.png
            loc = r"C:\\Users\\User\\downloads\\" + filenm
            # 사진 파일 복사해서 저장 - py 파일 경로에 temp.png 파일 생성
            shutil.copyfile(loc, 'temp.png')
        except:
            try:
                # 사진 복사 에러 시 진입 - 다운로드 시간 부여
                print('10초 프로토콜')
                # 대기
                time.sleep(10)
                # 사진 파일 경로 설정 - 다운로드 폴더 내 파일명.png
                loc = r"C:\\Users\\User\\downloads\\" + filenm
                # 사진 파일 복사해서 저장 - py 파일 경로에 temp.png 파일 생성
                shutil.copyfile(loc, 'temp.png')
            except:
                pass

        # 대기
        time.sleep(2)
        # 사진 파일 cv2로 최초 로딩
        image = cv2.imread("temp.png", 0)
        # BRG to RGB 로 형식 변경
        gray = cv2.cvtColor(image, cv2.COLOR_BGR2RGB)
        # 로딩 끝났으니 temp.png 파일 제거
        os.remove("temp.png")
        # 복사 끝났으니 다운로드 받은 파일 제거
        os.remove(loc)

        # 사진 규격 변수 입력 - 높이, 너비, 채널
        h, w, c = gray.shape
        print('file name: ', filenm, "/", "file id:  ", fileid, "/", 'width:  ', w, "/" 'height: ', h, "/", 'channel:',
              c)

        # 워터마크 시작지점 설정 (우하단)
        # OCR 위해 실제 워터마크보다 조금 작을 수 있음
        garo = w - 215
        sero = h - 50

        # 워터마크 부분만 잘라내서 변수에 저장 - 범위를 타이트하게 잡음 (10, 12 등 수치)
        crop = gray[sero:h - 10, garo:w - 12]

        # 배열로 변환
        kernel = np.array([[0, -1, 0],
                           [-1, 5, -1],
                           [0, -1, 0]])

        # 위 array 적용한거 i_crop에 저장
        i_crop = cv2.filter2D(crop, -1, kernel)

        # 비트연산 - 부정(bitwise_not) 연산임 - 여기서는 색 반전 목적으로 사용
        r_crop = cv2.bitwise_not(i_crop)

        # cv2 모든 창 닫기 - 하지만 여기서는 창을 열지 않아서 의미 없음
        cv2.destroyAllWindows()

        # psm6이 젤 좋음(2019) / psm4도 괜찮음
        custom_oem_psm_config = r'--oem 1 --psm 4'
        # 테서렉트에 r_crop 이미지 입력
        tess = pytesseract.image_to_string(r_crop, config=custom_oem_psm_config)
        # 테서렉트string 출력
        print('tess', tess)
        # 대기
        time.sleep(1)
    except:
        print('사진인식오류-테서렉트 점검')


def count_pic():
    # 이 함수에서 쓰던 변수를 global로 사용
    global b_window
    global a_window
    global numb
    global jetek
    global tess

    # 윈도우 2개 이상인 경우임
    b_window = driver.window_handles[0]
    a_window = driver.window_handles[1]

    # 두 번째 창으로 스위치
    driver.switch_to.window(a_window)

    # 현재 URL - cu
    cu = driver.current_url
    # URL 출력
    print(cu)

    # 제목 추출
    title = driver.find_element_by_xpath("""/html/body/div[2]/div[3]/div[3]/table/tbody/tr[2]/td""").text
    # 날짜 추출
    date = driver.find_element_by_xpath("""/html/body/div[2]/div[3]/div[1]/table/tbody/tr[1]/td[2]""").text
    # 보낸사람 추출
    repo = driver.find_element_by_id("repo_emp").text
    # 보고사항 추출
    bogosh = driver.find_element_by_xpath("""/html/body/div[2]/div[3]/div[3]/table/tbody/tr[3]/td""").text
    # 특이사항 추출
    tesh = driver.find_element_by_xpath("""/html/body/div[2]/div[3]/div[3]/table/tbody/tr[4]/td""").text
    # 보낸사람을 기준으로 한 근무조 추출 - 사전 입력 자료 근거
    jys = jetek_info[repo][0]
    # 보낸사람을 기준으로 한 근무 요일 추출 - 사전 입력 자료 근거
    yo1 = jetek_info[repo][1]
    # 근무조에 따른 업무보고 첨부 사진 자료
    aia = {'주간': [10, 11, 12, 13, 14, 15, 16], '야간': [19, 20, 21, 22, 23, 00], '심야': [1, 2, 3, 4]}
    # 사진 첨부 시간
    ea_time = aia[jys]
    # 첨부될 사진 개수
    ea_pics = len(ea_time)
    # 근무조 (사진 개수 등) 및 근무요일 출력
    print(jys, '(', ea_time, ' : ', ea_pics, ' 개의 사진 필요)', yo1)
    # 프레임 스위치 - 첨부파일이 위치한 프레임으로
    driver.switch_to.frame("voovoUpload")
    # 모든 class, file_set 추출 - file ID 추출 목적
    thum = driver.find_elements_by_class_name("file_set")
    # 모든 file name 추출
    thum_fn = driver.find_elements_by_name("fileNm")
    # file_set 개수
    len_t = len(thum)
    # 정수로 변경 - len 이라 정수일 것 같지만 소수점이 나와서 걸어줌
    len_t = int(len_t)

    # range식 넣기 위해서 1뺌 (사실 잘 기억 안남)
    len_t = len_t - 1

    # 리스트 변수 생성 - file id
    li = list()
    # 리스트 변수 생성 - file name
    ln = list()
    # 리스트 변수 생성 - 날짜, 시간(tess 전체)
    filetime = list()

    # 리스트 변수 생성 - tess 중 시간 부분
    olta_pic_s = list()
    # 리스트 변수 생성 - tess 중 날짜 부분
    nalza_s = list()
    # 리스트 변수 생성 - 에러
    pic_er = list()

    for i in range(0, len_t):
        # 새 창 포커스
        driver.switch_to.window(a_window)
        # 프레임 포커스
        driver.switch_to.frame("voovoUpload")
        # file_set class 중 i번째(파일 개수)
        aa = thum[i]
        # filenm 중 i번째(파일 개수)
        bb = thum_fn[i]

        # fileid 속성 추출 (from aa)
        fileid = aa.get_attribute("fileid")
        # i번째 fn 텍스트 추출
        filenm = bb.text
        # fileid를 li 리스트에 저장
        li.append(fileid)
        # filenm을 ln 리스트에 저장
        ln.append(filenm)

        # file name 개수
        len_fn = len(filenm)

        # file name에서 오른쪽 3글자 추출
        pnj = filenm[len_fn - 3:len_fn]

        # 오른쪽 3글자 소문자로 변경(lower)
        # png 파일 검사
        if pnj.lower() == 'png':
            # OCR 시작 안내문 출력
            print('사진OCR시작')
            # imgcrop 함수 실행
            imgcrop(fileid, filenm)
            # imgcrop 함수 종료 - tess변수 리턴

            # tess 중 왼쪽에서 4글자 비교 (2020과)
            # 공백이 없다는 가정 필요
            if tess[0:4] == '2020':
                # 왼쪽에서 10번째까지가 날짜
                tess_space = tess.replace(" ", "")
                nalza = tess_space[0:10]
                # 11번째부터 19번째까지가 시간
                olta_pic = tess_space[10:18]
                # 시간 0:2
                olta_h = olta_pic[0:2]
                # 분 3:5
                olta_m = olta_pic[3:5]
                # 초 6:8
                olta_s = olta_pic[6:8]

                # 시 분 초 출력
                print(olta_h, olta_m, olta_s)
                # 50분이 넘어가는 사진 분기
                if int(olta_m) > 49:
                    # 50분이 넘어가며 23시인 사진 분기
                    if int(olta_h) == 23:
                        # 00시로 변경
                        olta_h = '00'
                    else:
                        # 50분이 넘어가며 23시가 아닌 경우
                        olta_h = int(olta_h) + 1
                    try:
                        ea_time.remove(int(olta_h))
                    except:
                        pass
                elif int(olta_m) < 11:
                    print(olta_h, '시')
                    try:
                        ea_time.remove(int(olta_h))
                    except:
                        pass
                else:
                    print('해당안됨')

                olta_pic_s.append(olta_pic)
                nalza_s.append(nalza)
                print(nalza)
                print(olta_pic)
                filetime.append(tess)

            else:
                print('사진인식에러')
                pass
        else:
            print('PNG파일아님')
            pass

    print(ea_time)
    l_ea_time = len(ea_time)

    if det == 'Y':

        # det(댓글) 달기 실행
        if int(l_ea_time) == 0:
            ea_time = ''
        else:
            print(ea_time, l_ea_time, '이거')
            pass

        cmts = driver.find_elements_by_name("commentDiv")
        l_cmts = len(cmts) - 1
        print('댓글 개수:', l_cmts)

        if l_cmts == 0:
            print('댓글 달기 시작 1')
            if l_ea_time > 0:
                print('누락 사진이 1개 이상')
                aah = ''

                for ml in ea_time:
                    aah = aah + str(ml) + '시 '

                neyong = str(
                    l_ea_time) + '개의 사진 누락' + '(' + aah + '/' + "인정범위 : 1. 신고도구에서 다운로드 받은 PNG 파일, 2. 우하단 워터마크 정상 표시, 3. 정시 전후 10분 이내 사진)"
                print(neyong)
                driver.find_element_by_id("textDiv").send_keys(neyong)
                time.sleep(3)
                btb = driver.find_elements_by_class_name("submitBtn")[0]
                btb.click()

                try:
                    # 얼러트 종료
                    time.sleep(2)
                    alert = driver.switch_to.alert
                    alert.accept()
                    time.sleep(2)
                except:
                    pass
            else:
                pass
        else:
            pass
    else:
        print('댓글X')
        pass

    ln = str(ln)
    filetime = str(filetime)

    # ws1['A%s' % numb] = numb - 1
    # ws1['B%s' % numb] = repo
    # ws1['C%s' % numb] = title
    # ws1['D%s' % numb] = date
    # ws1['E%s' % numb] = len_t
    # ws1['F%s' % numb] = l_ea_time
    # ws1['G%s' % numb] = str(ea_time)
    # ws1['H%s' % numb] = bogosh
    # ws1['I%s' % numb] = tesh
    # ws1['J%s' % numb] = ln
    # ws1['K%s' % numb] = filetime
    # ws1['L%s' % numb] = cu

    ws1.append([numb - 1, repo, title, date, len_t, l_ea_time, str(ea_time), bogosh, tesh, ln, filetime, cu])

    numb = numb + 1
    driver.switch_to.window(a_window)
    driver.close()

    wb.save(filename=file_name)


go_bogo()

clickman()
